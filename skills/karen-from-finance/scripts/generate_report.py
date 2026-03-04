#!/usr/bin/env python3
"""
generate_report.py — Karen From Finance
Reads expenses.json, filters by requester phone number, outputs an Excel report.

Usage:
    python3 generate_report.py --requester +61400000001 \
        [--data ~/.openclaw/workspace/karen-data/expenses.json] \
        [--out /tmp/]
"""

import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


DEFAULT_DATA = Path.home() / ".openclaw" / "workspace" / "karen-data" / "expenses.json"
DEFAULT_OUT = Path.home() / ".openclaw" / "workspace" / "karen-data"

HEADER = ["Date", "Vendor", "Description", "Amount (ex GST)", "GST", "Total", "Category"]
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FONT = Font(bold=True)
CURRENCY_FORMAT = '#,##0.00'


def load_expenses(data_path: Path, requester: str) -> list[dict]:
    if not data_path.exists():
        sys.exit(f"Data file not found: {data_path}")
    with open(data_path) as f:
        all_expenses = json.load(f)
    return [e for e in all_expenses if e.get("requester") == requester]


def date_range(expenses: list[dict]) -> tuple[str, str]:
    dates = sorted(e["date"] for e in expenses)
    return dates[0], dates[-1]


def build_workbook(expenses: list[dict]) -> tuple[openpyxl.Workbook, str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    # Header row
    ws.append(HEADER)
    for col_idx, _ in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    total_ex_gst = 0.0
    total_gst = 0.0
    total_amount = 0.0

    for expense in sorted(expenses, key=lambda e: e["date"]):
        date_str = datetime.strptime(expense["date"], "%Y-%m-%d").strftime("%d/%m/%Y")
        category = expense.get("category", "").replace("_", " ").title()
        row = [
            date_str,
            expense.get("vendor", ""),
            expense.get("description", ""),
            expense.get("amount_ex_gst", 0),
            expense.get("gst", 0),
            expense.get("total", 0),
            category,
        ]
        ws.append(row)
        total_ex_gst += expense.get("amount_ex_gst", 0)
        total_gst += expense.get("gst", 0)
        total_amount += expense.get("total", 0)

    # Totals row
    totals_row = ws.max_row + 1
    ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
    for col, val in [(4, total_ex_gst), (5, total_gst), (6, total_amount)]:
        cell = ws.cell(row=totals_row, column=col, value=val)
        cell.font = Font(bold=True)

    # Currency formatting on amount columns (rows 2 onward)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT

    # Auto-fit column widths
    for col_idx in range(1, len(HEADER) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    # Build filename
    if expenses:
        start, end = date_range(expenses)
        filename = f"expenses_{start}_to_{end}.xlsx"
    else:
        filename = f"expenses_{date.today()}.xlsx"

    return wb, filename


def main():
    parser = argparse.ArgumentParser(description="Generate expense report for a requester")
    parser.add_argument("--requester", required=True, help="E.164 phone number, e.g. +61400000001")
    parser.add_argument("--data", type=Path, default=DEFAULT_DATA, help="Path to expenses.json")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT, help="Output directory")
    args = parser.parse_args()

    expenses = load_expenses(args.data, args.requester)

    if not expenses:
        print(f"No expenses found for {args.requester}")
        sys.exit(0)

    wb, filename = build_workbook(expenses)

    args.out.mkdir(parents=True, exist_ok=True)
    out_path = args.out / filename
    wb.save(out_path)

    total = sum(e.get("total", 0) for e in expenses)
    start, end = date_range(expenses)
    print(f"Report saved: {out_path}")
    print(f"{len(expenses)} expenses | ${total:.2f} total | {start} to {end}")


if __name__ == "__main__":
    main()
